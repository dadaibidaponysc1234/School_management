from user_registration.models import (ScoreObtainedPerAssessment, ScorePerAssessmentInstance,
                    ScoreObtainedPerAssessment, ResultConfiguration,Result, 
                    GradingSystem, ExamScore, ContinuousAssessment,AnnualResult,
                    AnnualResultWeightConfig,StudentClass,Year,Department,
                    StudentSubjectRegistration, Term,)

from django.db.models import Prefetch
from django.shortcuts import get_object_or_404

from django.db.models import Q
# excel_export.py
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


def is_result_viewable(school, result_type='term'):
    """
    Check if result (term or annual) is viewable for the given school.
    """
    try:
        control = school.result_visibility_control
        if result_type == 'term':
            return control.term_result_open
        elif result_type == 'annual':
            return control.annual_result_open
    except:
        return False

########==========================================########
def update_score_obtained_per_assessment(registration, category):
    """
    Updates or creates the total score for a specific registration and assessment category.
    Only considers instance scores within the active term.
    """
    # Sum all scores for this registration and category
    instances = ScorePerAssessmentInstance.objects.filter(
        registration=registration,
        category=category
    )

    total_score = sum([instance.score for instance in instances])

    # Update or create the total score record
    obj, _ = ScoreObtainedPerAssessment.objects.update_or_create(
        registration=registration,
        category=category,
        defaults={'total_score': total_score}
    )

    return obj

########==========================================########
def compute_continuous_assessment(registration):
    school = registration.school
    result_config = getattr(school, 'result_configuration', None)
    if not result_config:
        return 0.0

    ca_max = result_config.total_ca_score

    scores = ScoreObtainedPerAssessment.objects.filter(registration=registration)
    raw_total = sum([s.total_score for s in scores])
    max_possible = sum([
        cat.number_of_times * cat.max_score_per_one
        for cat in school.assessment_categories.all()
    ])

    if max_possible == 0:
        return 0.0

    scaled_total = (raw_total / max_possible) * ca_max
    return round(scaled_total, 2)

########==========================================########
def compute_result_for_registration(registration):
    """
    Computes and stores a Result for a student's subject registration.
    Uses ContinuousAssessment and ExamScore values.
    """
    try:
        ca_obj = registration.continuous_assessments
        exam_obj = registration.exam_scores.first()
    except:
        return None

    if not ca_obj or not exam_obj:
        return None

    ca_total = ca_obj.ca_total
    exam_score = exam_obj.score
    total = ca_total + exam_score

    school = registration.school
    grading = GradingSystem.objects.filter(
        school=school, min_score__lte=total, max_score__gte=total
    ).first()

    grade = grading.grade if grading else 'N/A'
    remarks = grading.remarks if grading else 'Not Graded'

    result, _ = Result.objects.update_or_create(
        registration=registration,
        defaults={
            'ca_total': ca_total,
            'exam_score': exam_score,
            'total_score': total,
            'grade': grade,
            'remarks': remarks
        }
    )
    return result
########==========================================########

def compute_annual_result(registration):
    """
    Computes or updates the AnnualResult for a given StudentSubjectRegistration.
    Looks up Result records for each term, applies weight configuration if present,
    otherwise falls back to rules:
      - 3 terms: equal weight
      - 2 terms: 50% each
      - 1 term: must be 3rd term (100%)
    """
    # Fetch term result scores
    term_results = Result.objects.filter(
        registration__student_class=registration.student_class,
        registration__subject_class=registration.subject_class,
        registration__student_class__student=registration.student_class.student
    )

    first_term_result = term_results.filter(registration__term__name__iexact='First Term').first()
    second_term_result = term_results.filter(registration__term__name__iexact='Second Term').first()
    third_term_result = term_results.filter(registration__term__name__iexact='Third Term').first()

    first_score = first_term_result.total_score if first_term_result else None
    second_score = second_term_result.total_score if second_term_result else None
    third_score = third_term_result.total_score if third_term_result else None

    # Check configured weights
    weight_config = AnnualResultWeightConfig.objects.filter(
        school=registration.school,
        class_year=registration.student_class.class_year,
        department=registration.subject_class.department
    ).first()

    if weight_config:
        annual_average = (
            (first_score or 0) * weight_config.first_term_weight +
            (second_score or 0) * weight_config.second_term_weight +
            (third_score or 0) * weight_config.third_term_weight
        )
    else:
        # Apply fallback logic
        present_scores = {
            'first': first_score,
            'second': second_score,
            'third': third_score
        }
        available = {k: v for k, v in present_scores.items() if v is not None}
        count = len(available)

        if count == 3:
            annual_average = (first_score + second_score + third_score) / 3
        elif count == 2:
            annual_average = sum(available.values()) / 2
        elif count == 1 and 'third' in available:
            annual_average = third_score
        else:
            # Not enough valid data
            return None

    # Fetch grade
    grading = GradingSystem.objects.filter(
        school=registration.school,
        min_score__lte=annual_average,
        max_score__gte=annual_average
    ).first()

    grade = grading.grade if grading else 'N/A'
    remarks = grading.remarks if grading else 'Not Graded'

    # Save or update record
    annual_result, _ = AnnualResult.objects.update_or_create(
        registration=registration,
        defaults={
            'first_term_score': first_score,
            'second_term_score': second_score,
            'third_term_score': third_score,
            'annual_average': round(annual_average, 2),
            'grade': grade,
            'remarks': remarks
        }
    )

    return annual_result
################################### Full Result#########################
# utils.py

def calculate_grade(score, school):
    grading = GradingSystem.objects.filter(school=school).order_by('-max_score')
    for g in grading:
        if g.min_score <= score <= g.max_score:
            return g.grade, g.remarks
    return "", ""

def get_full_term_result_data(student, year_id, term):
    year = get_object_or_404(Year, year_id=year_id)
    results = Result.objects.filter(
        registration__student_class__student=student,
        registration__term=term,
        registration__term__year=year
    ).select_related('registration__subject_class__subject', 'registration')

    data = {
        "school": get_school_info(student.school),
        "student": get_student_info(student),
        "year": year.name,
        "term": term.name,
        "term_results": []
    }

    for result in results:
        reg = result.registration
        subject = reg.subject_class.subject
        assessments = []
        assessment_scores = ScoreObtainedPerAssessment.objects.filter(registration=reg)
        for a in assessment_scores:
            max_score = a.category.number_of_times * a.category.max_score_per_one
            assessments.append({
                "assessment_name": a.category.assessment_name,
                "obtained_score": a.total_score,
                "max_score": max_score
            })

        data["term_results"].append({
            "subject": {"subject_id": str(subject.subject_id), "name": subject.name},
            "class_year": reg.class_year_name,
            "class_arm": reg.class_arm_name,
            "ca_total": result.ca_total,
            "exam_score": result.exam_score,
            "total_score": result.total_score,
            "grade": result.grade,
            "remarks": result.remarks,
            "assessments": assessments
        })
    return data

def get_full_annual_result_data(student, year_id):
    year = get_object_or_404(Year, year_id=year_id)
    annuals = AnnualResult.objects.filter(
        registration__student_class__student=student,
        registration__term__year=year
    ).select_related('registration__subject_class__subject', 'registration')

    school = student.school
    data = {
        "school": get_school_info(school),
        "student": get_student_info(student),
        "year": year.name,
        "annual_results": []
    }

    for annual in annuals:
        reg = annual.registration
        subject = reg.subject_class.subject

        config = AnnualResultWeightConfig.objects.filter(
            school=school,
            class_year=reg.student_class.class_year,
            department=reg.subject_class.department
        ).first()

        f, s, t = annual.first_term_score or 0, annual.second_term_score or 0, annual.third_term_score or 0
        if config:
            weighted = {
                "first_term": round(f * config.first_term_weight, 2),
                "second_term": round(s * config.second_term_weight, 2),
                "third_term": round(t * config.third_term_weight, 2),
            }
        else:
            terms_present = [score for score in [f, s, t] if score > 0]
            if len(terms_present) == 3:
                weighted = {"first_term": f/3, "second_term": s/3, "third_term": t/3}
            elif len(terms_present) == 2:
                weighted = {"first_term": f/2, "second_term": s/2, "third_term": t/2}
            else:
                weighted = {"first_term": 0, "second_term": 0, "third_term": t} if t else {"first_term": f, "second_term": 0, "third_term": 0}

        avg = round(sum(weighted.values()), 2)
        grade, remarks = calculate_grade(avg, school)

        third_ca = ContinuousAssessment.objects.filter(registration=reg).first()
        third_exam = ExamScore.objects.filter(registration=reg).first()
        third_assessments = []
        if third_ca:
            scored_assessments = ScoreObtainedPerAssessment.objects.filter(registration=reg)
            for a in scored_assessments:
                max_score = a.category.number_of_times * a.category.max_score_per_one
                third_assessments.append({
                    "assessment_name": a.category.assessment_name,
                    "obtained_score": a.total_score,
                    "max_score": max_score
                })

        data["annual_results"].append({
            "subject": {"subject_id": str(subject.subject_id), "name": subject.name},
            "class_year": reg.class_year_name,
            "class_arm": reg.class_arm_name,
            "weights_used": {
                "first_term_weight": config.first_term_weight if config else None,
                "second_term_weight": config.second_term_weight if config else None,
                "third_term_weight": config.third_term_weight if config else None
            },
            "weighted_term_scores": weighted,
            "annual_average": avg,
            "grade": grade,
            "remarks": remarks,
            "third_term_assessments": {
                "ca_total": third_ca.ca_total if third_ca else None,
                "exam_score": third_exam.score if third_exam else None,
                "assessments": third_assessments
            }
        })
    return data

def get_school_info(school):
    return {
        "school_id": str(school.id),
        "school_name": school.school_name,
        "short_name": school.short_name,
        "school_type": school.school_type,
        "education_level": school.education_level,
        "city": school.city,
        "state": school.state,
        "country": school.country,
        "logo_url": school.logo.url if school.logo else None
    }

def get_student_info(student):
    return {
        "student_id": str(student.student_id),
        "first_name": student.first_name,
        "last_name": student.last_name,
        "gender": student.gender,
        "admission_number": student.admission_number
    }
################################ Broadsheet #####################
# utils_broadsheet.py
def calculate_grade(score, school):
    grading = GradingSystem.objects.filter(school=school).order_by('-max_score')
    for g in grading:
        if g.min_score <= score <= g.max_score:
            return g.grade
    return "F"


def calculate_position(students, key):
    sorted_students = sorted(
        [s for s in students if s.get(key) is not None],
        key=lambda x: x[key], reverse=True
    )
    positions = {}
    for i, s in enumerate(sorted_students):
        previous = sorted_students[i - 1][key] if i > 0 else None
        if previous == s[key]:
            positions[s['student_id']] = positions[sorted_students[i - 1]['student_id']]
        else:
            positions[s['student_id']] = f"{i + 1}{ordinal_suffix(i + 1)}"
    return positions


def ordinal_suffix(n):
    return "th" if 11 <= n % 100 <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")


def get_broadsheet_data(class_year=None, class_arm=None, year=None, term=None):
    school = class_year.school if class_year else class_arm.school
    department = class_year.department if class_year else class_arm.department
    is_primary_or_junior = department.name.lower() in ["junior", "primary"]

    result_config = ResultConfiguration.objects.filter(school=school).first()
    pass_mark = result_config.pass_mark if result_config else 50

    # filter registrations
    filters = {
        "school": school,
        "term__year": year
    }
    if term:
        filters["term"] = term
    if class_year:
        filters["student_class__class_year"] = class_year
    if class_arm:
        filters["student_class__class_arm"] = class_arm

    registrations = StudentSubjectRegistration.objects.filter(**filters)

    subjects = list({reg.subject_class.subject.name for reg in registrations})
    students = list({reg.student_class.student for reg in registrations})

    broadsheet = {
        "broadsheet_type": "term" if term else "annual",
        "year": year.name,
        "term": term.name if term else None,
        "class_year": class_year.class_name if class_year else None,
        "class_arm": class_arm.classes.arm_name if class_arm else None,
        "department": department.name,
        "pass_mark": pass_mark,
        "subjects": subjects,
        "students": []
    }

    student_rows = []

    for student in students:
        row = {
            "student_id": str(student.student_id),
            "name": f"{student.first_name} {student.last_name}",
            "scores": {},
            "passed_subjects": 0,
            "failed_subjects": 0
        }
        subject_scores = []

        for subject in subjects:
            if term:
                result = Result.objects.filter(
                    registration__student_class__student=student,
                    registration__term=term,
                    registration__subject_class__subject__name=subject
                ).first()
                if result:
                    score = result.total_score
                    grade = result.grade or calculate_grade(score, school)
                    row["scores"][subject] = {"score": score, "grade": grade}
                    subject_scores.append(score)
                    if score >= pass_mark:
                        row["passed_subjects"] += 1
                    else:
                        row["failed_subjects"] += 1
            else:
                annual_result = AnnualResult.objects.filter(
                    registration__student_class__student=student,
                    registration__term__year=year,
                    registration__subject_class__subject__name=subject
                ).first()
                if annual_result:
                    avg = annual_result.annual_average
                    grade = annual_result.grade or calculate_grade(avg, school)
                    row["scores"][subject] = {"annual_average": avg, "grade": grade}
                    subject_scores.append(avg)
                    if avg >= pass_mark:
                        row["passed_subjects"] += 1
                    else:
                        row["failed_subjects"] += 1

        if is_primary_or_junior and subject_scores:
            row["average_score"] = round(sum(subject_scores) / len(subject_scores), 2)

        student_rows.append(row)

    if is_primary_or_junior:
        avg_positions = calculate_position(student_rows, "average_score")
        for row in student_rows:
            row["position_in_class_year"] = avg_positions.get(row["student_id"])
            row["position_in_class_arm"] = avg_positions.get(row["student_id"])

    broadsheet["students"] = student_rows
    return broadsheet
####################################################################

def export_broadsheet_to_excel(broadsheet_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Broadsheet"

    # Title
    title = f"{broadsheet_data['class_year']} {broadsheet_data['class_arm']} Broadsheet - {broadsheet_data['year']}"
    if broadsheet_data["broadsheet_type"] == "term":
        title += f" ({broadsheet_data['term']})"
    ws.merge_cells('A1:H1')
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Student Name"] + broadsheet_data["subjects"]
    if "average_score" in broadsheet_data["students"][0]:
        headers += ["Average", "Class Pos", "Arm Pos"]
    headers += ["Passed", "Failed"]

    ws.append(headers)

    for student in broadsheet_data["students"]:
        row = [student["name"]]
        for subject in broadsheet_data["subjects"]:
            subj_data = student["scores"].get(subject, {})
            score = subj_data.get("score") or subj_data.get("annual_average")
            grade = subj_data.get("grade")
            row.append(f"{score} ({grade})" if score is not None else "-")

        if "average_score" in student:
            row.append(student.get("average_score", "-"))
            row.append(student.get("position_in_class_year", "-"))
            row.append(student.get("position_in_class_arm", "-"))

        row.append(student.get("passed_subjects", 0))
        row.append(student.get("failed_subjects", 0))

        ws.append(row)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"broadsheet_{broadsheet_data['class_year']}_{broadsheet_data['class_arm']}_{broadsheet_data['year']}"
    if broadsheet_data['broadsheet_type'] == "term":
        filename += f"_{broadsheet_data['term'].replace(' ', '_')}"
    filename += ".xlsx"

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



